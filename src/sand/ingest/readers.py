"""Path-based spreadsheet readers (legacy Excel .xls helpers)."""

from __future__ import annotations

from pathlib import Path

import pandas as pd

SUPPORTED_EXTENSIONS = {".csv", ".xlsx", ".xls", ".parquet"}


def list_xlsx_sheets(path: str | Path) -> list[str]:
    """List sheet names without loading cell data into pandas."""
    from openpyxl import load_workbook

    wb = load_workbook(Path(path), read_only=True, data_only=True)
    try:
        return list(wb.sheetnames)
    finally:
        wb.close()


def read_spreadsheet(path: str | Path) -> dict[str, pd.DataFrame]:
    """Read a spreadsheet into ``{sheet_name: DataFrame}``.

    Prefer DuckDB native ingest for CSV/Parquet/XLSX. This path remains for
    legacy ``.xls`` and callers that still want pandas sheets.
    """
    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(f"File not found: {path}")

    ext = path.suffix.lower()
    if ext not in SUPPORTED_EXTENSIONS:
        raise ValueError(f"Unsupported file type '{ext}'. Use: {', '.join(sorted(SUPPORTED_EXTENSIONS))}")

    if ext == ".csv":
        df = pd.read_csv(path)
        return {path.stem: _normalize_columns(df)}

    if ext == ".parquet":
        df = pd.read_parquet(path)
        return {path.stem: _normalize_columns(df)}

    sheets = pd.read_excel(path, sheet_name=None, engine="openpyxl" if ext == ".xlsx" else None)
    return {name: _normalize_columns(df) for name, df in sheets.items()}


def _normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    out.columns = [str(c).strip() if c is not None else f"col_{i}" for i, c in enumerate(out.columns)]
    return out
